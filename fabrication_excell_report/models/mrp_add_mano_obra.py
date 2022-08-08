# -*- coding: utf-8 -*-
from odoo.tools.misc import DEFAULT_SERVER_DATETIME_FORMAT
import time
import odoo.addons.decimal_precision as dp
from openerp.osv import osv
import base64
from odoo import models, fields, api
import codecs

from datetime import datetime, timedelta
def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
values = {}
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
values = {}

def border(ws,texto):
	cell = WriteOnlyCell(ws, value=texto)
	cell.font = Font(name='Courier',size=14,bold=True)
	cell.border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000'))
	return cell

class mrp_production(models.Model):
	_inherit = "mrp.production"

	def do_csvtoexcel(self):




		cad = ""
		s_prod = [-1,-1,-1]
		s_loca = [-1,-1,-1]
	
		import io
		output = io.BytesIO()

		workbook = Workbook(write_only=True)
		ws = workbook.create_sheet("Reporte Orden De Producción")
		x= 9
		tam_col = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        boldbord = workbook.add_format({'bold': True})
		boldbord.set_align('center')
		boldbord.set_align('vcenter')
		boldbord.set_text_wrap()
		boldbord.font = Font(name='Calibri',size=11,bold=True)
        boldbord.set_font_size(11)



		cell = WriteOnlyCell(ws, value="FORMATO DE ORDEN DE PRODUCCIÓN - STOCK")
		cell.font = Font(name='Calibri',size=18,bold=True)
        cell.set_bg_color('#DCE6F1')
		cell.alignment = Alignment(horizontal='center')
        
		ws.merged_cells.ranges.append(get_column_letter(4)+ "1:" + get_column_letter(11) + '1')
		ws.append(["","","","","","",cell])

		ws.append([""])
        ws.write(1,4, "CONCEPTO",boldbord)
        ws.write(1,5, "Camb. Codigo",boldbord)
        ws.write(1,6, "Mezcla",boldbord)
        ws.write(1,7, "Reenvase",boldbord)
        ws.write(1,8, "Dilucion",boldbord)
        ws.write(4,5, "FECHA SOL:",boldbord)
        ws.write(5,5, str(self.date_planned_start),boldbord)
        ws.write(4,6, "FECHA PROD:",boldbord)
        ws.write(5,6, str(self.date_planned_start),boldbord)
        
		ws.append([""])
		ws.append([""])
		ws.append([""])

			
		import datetime		


		import datetime

		workbook.save(output)
		output.seek(0)

		attach_id = self.env['ir.attachment'].create({
                    'name': "FORMATO DE ORDEN DE PRODUCCIÓN.xlsx",
                    'type': 'binary',
                    'datas': base64.encodestring(output.read()),
                    'eliminar_automatico': True
                })
		output.close()


		return {
			'notif_button':{'auto_close':False,'with_menssage':1,'title':'FORMATO DE ORDEN DE PRODUCCIÓN','message':'Se proceso ','eventID':attach_id.id,'model_notify':'ir.attachment','method_notify':'get_download_ls','name_button':'Descargar FORMATO DE ORDEN DE PRODUCCIÓN'}
		}
	
